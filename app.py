import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Настройки страницы
st.set_page_config(
    page_title="Прайс КРАЙВИН 2026",
    page_icon="🍷",
    layout="wide"
)

st.title("🍷 Прайс-Калькулятор КРАЙВИН 2026")
st.caption("Расчет цен для клиентов, минимальной цены и цены на полке")

st.divider()

# База предустановленных товаров
PRESET_PRODUCTS = {
    "Мезенка 24": {"year": 2024, "price_disc": 531.0, "m_buy": 25.1, "m_tk": 10.0, "shelf": 1190.0},
    "Татуаж 24": {"year": 2024, "price_disc": 441.0, "m_buy": 25.1, "m_tk": 10.0, "shelf": 990.0},
    "Уикенд 24": {"year": 2024, "price_disc": 441.0, "m_buy": 25.1, "m_tk": 10.0, "shelf": 990.0},
    "Катарон 25": {"year": 2025, "price_disc": 800.0, "m_buy": 25.0, "m_tk": 5.0, "shelf": 1600.0},
    "Шарма 25": {"year": 2025, "price_disc": 650.0, "m_buy": 25.1, "m_tk": 5.0, "shelf": 1300.0},
    "Мезенка 25": {"year": 2025, "price_disc": 650.0, "m_buy": 25.1, "m_tk": 5.0, "shelf": 1300.0}
}

# Функция для генерации красивого Excel с границами
def create_styled_excel(df, sheet_name="Сводный"):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        ws = writer.sheets[sheet_name]
        
        # Стили
        thin_border = Border(
            left=Side(style='thin', color='B0BEC5'),
            right=Side(style='thin', color='B0BEC5'),
            top=Side(style='thin', color='B0BEC5'),
            bottom=Side(style='thin', color='B0BEC5')
        )
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        data_font = Font(name='Calibri', size=11)
        
        # Настройка ширин колонок и границ
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            for cell in col:
                cell.border = thin_border
                cell.font = data_font
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    return buffer.getvalue()


# Переключатель режимов
app_mode = st.radio(
    "Выберите режим работы:",
    ["🧮 Интерактивный калькулятор (Мультивыбор)", "📁 Расчет из файла Excel"],
    horizontal=True
)

st.divider()

# Боковая панель
st.sidebar.header("⚙️ Глобальные наценки (по умолчанию)")
default_markup_buy = st.sidebar.number_input("Наценка на закупку (%)", min_value=0.0, max_value=100.0, value=25.0, step=0.5)
default_markup_tk = st.sidebar.number_input("Наценка на подвоз до ТК (%)", min_value=0.0, max_value=50.0, value=5.0, step=1.0)


# --- 1. ИНТЕРАКТИВНЫЙ КАЛЬКУЛЯТОР С МУЛЬТИВЫБОРОМ ---
if "Мультивыбор" in app_mode:

    st.subheader("1. Выберите позиции из каталога")

    selected_products = st.multiselect(
        "Выберите одну или несколько серий:",
        options=list(PRESET_PRODUCTS.keys()),
        default=list(PRESET_PRODUCTS.keys())
    )

    st.divider()

    if selected_products:
        st.subheader("2. Корректировка параметров и расчёт")
        
        calculated_rows = []

        for prod_name in selected_products:
            preset = PRESET_PRODUCTS[prod_name]
            
            with st.expander(f"⚙️ Параметры: **{prod_name}**", expanded=False):
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    year_v = st.number_input(f"ГОД ({prod_name})", min_value=2024, max_value=2030, value=preset["year"], key=f"y_{prod_name}")
                with col2:
                    price_d = st.number_input(f"Цена со скидкой (руб) ({prod_name})", min_value=1.0, value=preset["price_disc"], key=f"p_{prod_name}")
                with col3:
                    m_b = st.number_input(f"Наценка закупки (%) ({prod_name})", min_value=0.0, value=preset["m_buy"], key=f"mb_{prod_name}") / 100.0
                with col4:
                    m_t = st.number_input(f"Наценка ТК (%) ({prod_name})", min_value=0.0, value=preset["m_tk"], key=f"mt_{prod_name}") / 100.0
                
                shelf_p = st.number_input(f"Цена на полке (руб) ({prod_name})", min_value=1.0, value=preset["shelf"], key=f"sh_{prod_name}")

            # Формулы
            kravin_in = price_d
            moscow_price = round(kravin_in * (1.0 + m_b), 2)
            min_price = round(moscow_price * (1.0 + m_t), 2)
            shelf_markup_pct = round(((shelf_p - min_price) / min_price) * 100.0, 1) if min_price > 0 else 0.0

            calculated_rows.append({
                "Серия": prod_name,
                "ГОД": year_v,
                "Цена со скидкой 10%": price_d,
                "Входная цена КРАЙВИН (руб/бут)": kravin_in,
                "Наценка на закупку": f"{m_b*100:.1f}%",
                "Цена из Москвы для клиентов (руб/бут)": moscow_price,
                "Наценка на подвоз до ТК": f"{m_t*100:.1f}%",
                "Минимальная цена (руб/бут)": min_price,
                "Рекомендуемая цена на полке (руб/бут)": shelf_p,
                "% наценки от минимальной цены": f"{shelf_markup_pct}%"
            })

        st.subheader("3. Сводная таблица выбранных позиций")
        df_calc = pd.DataFrame(calculated_rows)
        st.dataframe(df_calc, use_container_width=True, hide_index=True)

        # Выгрузка с границами
        excel_data = create_styled_excel(df_calc, sheet_name="Сводный")

        st.download_button(
            label="📥 Скачать файл «Сводный.xlsx»",
            data=excel_data,
            file_name="Сводный.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Пожалуйста, выберите хотя бы одну позицию из списка выше.")


# --- 2. РАСЧЕТ ИЗ ФАЙЛА EXCEL ---
else:
    st.subheader("📁 Автоматический расчет из Excel")
    uploaded_file = st.file_uploader("Загрузить файл Excel (.xlsx)", type=["xlsx", "xls"])

    if uploaded_file is not None:
        try:
            df_raw = pd.read_excel(uploaded_file, header=1)
            st.success("Файл успешно загружен!")

            results = []
            for idx, row in df_raw.iterrows():
                series = row.get("Серия", row.iloc[1] if len(row) > 1 else None)
                
                if pd.isna(series) or str(series).startswith("Цена включает") or str(series).startswith("*"):
                    continue

                year = int(row.get("ГОД", 2025)) if not pd.isna(row.get("ГОД")) else 2025
                price_disc = float(row.get("Цена со скидкой 10% на 2024 г.", row.iloc[3] if len(row) > 3 else 0))
                
                if price_disc > 0:
                    raw_m_buy = float(row.get("Наценка на закупку", default_markup_buy / 100.0))
                    m_buy = raw_m_buy / 100.0 if raw_m_buy > 1.0 else raw_m_buy

                    raw_m_tk = float(row.get("Наценка на подвоз до ТК", default_markup_tk / 100.0))
                    m_tk = raw_m_tk / 100.0 if raw_m_tk > 1.0 else raw_m_tk

                    shelf = float(row.get("Рекомендуемая цена на полке", row.iloc[9] if len(row) > 9 else 0))

                    k_in = price_disc
                    p_moscow = round(k_in * (1.0 + m_buy), 2)
                    p_min = round(p_moscow * (1.0 + m_tk), 2)
                    pct_shelf = round(((shelf - p_min) / p_min) * 100.0, 1) if p_min > 0 else 0.0

                    results.append({
                        "Серия": series,
                        "ГОД": year,
                        "Цена со скидкой 10%": round(price_disc, 2),
                        "Входная цена КРАЙВИН (руб/бут)": round(k_in, 2),
                        "Наценка на закупку": f"{m_buy * 100:.1f}%",
                        "Цена из Москвы для клиентов (кроме ЮФО) (руб/бут)": p_moscow,
                        "Наценка на подвоз до ТК": f"{m_tk * 100:.1f}%",
                        "Минимальная цена (руб/бут)": p_min,
                        "Рекомендуемая цена на полке": round(shelf, 2),
                        "% наценки от минимальной цены": f"{pct_shelf}%"
                    })

            if results:
                df_res = pd.DataFrame(results)
                st.dataframe(df_res, use_container_width=True, hide_index=True)

                # Выгрузка с границами
                excel_data_file = create_styled_excel(df_res, sheet_name="Сводный")

                st.download_button(
                    label="📥 Скачать файл «Сводный.xlsx»",
                    data=excel_data_file,
                    file_name="Сводный.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Ошибка при обработке файла: {e}")
